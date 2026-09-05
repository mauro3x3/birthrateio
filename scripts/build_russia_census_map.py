#!/usr/bin/env python3
"""Build a Russia ethnicity choropleth from Rosstat VPN-2020 Volume 5 Table 1.

Source: national composition of the population by federal subject
(urban+rural, both sexes). Shares are of total resident population,
including people who did not state an ethnicity.
"""

from __future__ import annotations

import json
import sys
import urllib.request
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT_GEO = ROOT / "public" / "geo" / "census"
OUT_DATA = ROOT / "public" / "data" / "census"
OUT_CATALOG = ROOT / "src" / "lib" / "data" / "census-maps-catalog.json"
GEO_SRC = ROOT / "public" / "geo" / "maps" / "rus-tfr.json"
CACHE = ROOT / ".cache" / "census"
XLSX_PATH = CACHE / "rosstat-vpn2020-tab1.xlsx"
XLSX_URL = (
    "https://web.archive.org/web/20231213104033/"
    "https://rosstat.gov.ru/storage/mediabank/Tom5_tab1_VPN-2020.xlsx"
)

SOURCE = (
    "Rosstat — 2020 All-Russian Population Census (VPN-2020), "
    "Volume 5 Table 1, national composition by federal subject"
)
SOURCE_URL = "https://rosstat.gov.ru/vpn/2020/Tom5_Nacionalnyj_sostav_i_vladenie_yazykami"

GROUPS = [
    {
        "id": "russian",
        "shortLabel": "Russians",
        "label": "Ethnic Russians (including Cossacks and Pomors)",
    },
    {
        "id": "tatar",
        "shortLabel": "Tatars",
        "label": "Tatars (excluding Crimean Tatars)",
    },
    {
        "id": "chechen",
        "shortLabel": "Chechens",
        "label": "Chechens",
    },
    {
        "id": "bashkir",
        "shortLabel": "Bashkirs",
        "label": "Bashkirs",
    },
    {
        "id": "other",
        "shortLabel": "Other",
        "label": "All other stated ethnic groups",
    },
    {
        "id": "not_stated",
        "shortLabel": "Not stated",
        "label": "Ethnicity not stated on the census form",
    },
]

# Rosstat sheet title → existing map slug. Skip RF, and the Tyumen /
# Arkhangelsk sheets that include autonomous okrugs (those okrugs are
# separate polygons).
SHEET_TO_SLUG = {
    "Белгородская область": "russia-belgorod-oblast",
    "Брянская область": "russia-bryansk-oblast",
    "Владимирская область": "russia-vladimir-oblast",
    "Воронежская область": "russia-voronezh-oblast",
    "Ивановская область": "russia-ivanovo-oblast",
    "Калужская область": "russia-kaluga-oblast",
    "Костромская область": "russia-kostroma-oblast",
    "Курская область": "russia-kursk-oblast",
    "Липецкая область": "russia-lipetsk-oblast",
    "Московская область": "russia-moscow-oblast",
    "Орловская область": "russia-oryol-oblast",
    "Рязанская область": "russia-ryazan-oblast",
    "Смоленская область": "russia-smolensk-oblast",
    "Тамбовская область": "russia-tambov-oblast",
    "Тверская область": "russia-tver-oblast",
    "Тульская область": "russia-tula-oblast",
    "Ярославская область": "russia-yaroslavl-oblast",
    "г. Москва": "russia-moscow",
    "Республика Карелия": "russia-karelia",
    "Республика Коми": "russia-komi",
    "Архангельская область без АО": "russia-arkhangelsk-oblast",
    "Ненецкий автономный округ": "russia-nenets",
    "Вологодская область": "russia-vologda-oblast",
    "Калининградская область": "russia-kaliningrad-oblast",
    "Ленинградская область": "russia-leningrad-oblast",
    "Мурманская область": "russia-murmansk-oblast",
    "Новгородская область": "russia-novgorod-oblast",
    "Псковская область": "russia-pskov-oblast",
    "г. Санкт-Петербург": "russia-saint-petersburg",
    "Республика Адыгея": "russia-adygea",
    "Республика Калмыкия": "russia-kalmykia",
    "Республика Крым": "russia-crimea",
    "Краснодарский край": "russia-krasnodar-krai",
    "Астраханская область": "russia-astrakhan-oblast",
    "Волгоградская область": "russia-volgograd-oblast",
    "Ростовская область": "russia-rostov-oblast",
    "г. Севастополь": "russia-sevastopol",
    "Республика Дагестан": "russia-dagestan",
    "Республика Ингушетия": "russia-ingushetia",
    "Кабардино-Балкарская Республика": "russia-kabardino-balkaria",
    "Карачаево-Черкесская Республика": "russia-karachay-cherkessia",
    "РСО-Алания": "russia-north-ossetia",
    "Чеченская Республика": "russia-chechnya",
    "Ставропольский край": "russia-stavropol-krai",
    "Республика Башкортостан": "russia-bashkortostan",
    "Республика Марий Эл": "russia-mari-el",
    "Республика Мордовия": "russia-mordovia",
    "Республика Татарстан": "russia-tatarstan",
    "Удмуртская Республика": "russia-udmurtia",
    "Чувашская Республика": "russia-chuvashia",
    "Пермский край": "russia-perm-krai",
    "Кировская область": "russia-kirov-oblast",
    "Нижегородская область": "russia-nizhny-novgorod-oblast",
    "Оренбургская область": "russia-orenburg-oblast",
    "Пензенская область": "russia-penza-oblast",
    "Самарская область": "russia-samara-oblast",
    "Саратовская область": "russia-saratov-oblast",
    "Ульяновская область": "russia-ulyanovsk-oblast",
    "Курганская область": "russia-kurgan-oblast",
    "Свердловская область": "russia-sverdlovsk-oblast",
    "Тюменская область без АО": "russia-tyumen-oblast",
    "ХМАО": "russia-khanty-mansi",
    "ЯНАО": "russia-yamalo-nenets",
    "Челябинская область": "russia-chelyabinsk-oblast",
    "Республика Алтай": "russia-altai-republic",
    "Республика Тыва": "russia-tuva",
    "Республика Хакасия": "russia-khakassia",
    "Алтайский край": "russia-altai-krai",
    "Красноярский край": "russia-krasnoyarsk-krai",
    "Иркутская область": "russia-irkutsk-oblast",
    "Кемеровская область - Кузбасс": "russia-kemerovo-oblast",
    "Новосибирская область": "russia-novosibirsk-oblast",
    "Омская область": "russia-omsk-oblast",
    "Томская область": "russia-tomsk-oblast",
    "Республика Бурятия": "russia-buryatia",
    "Республика Саха (Якутия)": "russia-sakha",
    "Забайкальский край": "russia-zabaykalsky-krai",
    "Камчатский край": "russia-kamchatka-krai",
    "Приморский край": "russia-primorsky-krai",
    "Хабаровский край": "russia-khabarovsk-krai",
    "Амурская область": "russia-amur-oblast",
    "Магаданская область": "russia-magadan-oblast",
    "Сахалинская область": "russia-sakhalin-oblast",
    "Еврейская автономная область": "russia-jewish-autonomous-oblast",
    "Чукотский автономный округ": "russia-chukotka",
}

DISPLAY_NAMES = {
    "russia-adygea": "Republic of Adygea",
    "russia-altai-krai": "Altai Krai",
    "russia-altai-republic": "Altai Republic",
    "russia-amur-oblast": "Amur Oblast",
    "russia-arkhangelsk-oblast": "Arkhangelsk Oblast",
    "russia-astrakhan-oblast": "Astrakhan Oblast",
    "russia-bashkortostan": "Republic of Bashkortostan",
    "russia-belgorod-oblast": "Belgorod Oblast",
    "russia-bryansk-oblast": "Bryansk Oblast",
    "russia-buryatia": "Republic of Buryatia",
    "russia-chechnya": "Chechen Republic",
    "russia-chelyabinsk-oblast": "Chelyabinsk Oblast",
    "russia-chukotka": "Chukotka Autonomous Okrug",
    "russia-chuvashia": "Chuvash Republic",
    "russia-crimea": "Republic of Crimea",
    "russia-dagestan": "Republic of Dagestan",
    "russia-ingushetia": "Republic of Ingushetia",
    "russia-irkutsk-oblast": "Irkutsk Oblast",
    "russia-ivanovo-oblast": "Ivanovo Oblast",
    "russia-jewish-autonomous-oblast": "Jewish Autonomous Oblast",
    "russia-kabardino-balkaria": "Kabardino-Balkaria",
    "russia-kaliningrad-oblast": "Kaliningrad Oblast",
    "russia-kalmykia": "Republic of Kalmykia",
    "russia-kaluga-oblast": "Kaluga Oblast",
    "russia-kamchatka-krai": "Kamchatka Krai",
    "russia-karachay-cherkessia": "Karachay-Cherkess Republic",
    "russia-karelia": "Republic of Karelia",
    "russia-kemerovo-oblast": "Kemerovo Oblast",
    "russia-khabarovsk-krai": "Khabarovsk Krai",
    "russia-khakassia": "Republic of Khakassia",
    "russia-khanty-mansi": "Khanty-Mansi Autonomous Okrug",
    "russia-kirov-oblast": "Kirov Oblast",
    "russia-komi": "Komi Republic",
    "russia-kostroma-oblast": "Kostroma Oblast",
    "russia-krasnodar-krai": "Krasnodar Krai",
    "russia-krasnoyarsk-krai": "Krasnoyarsk Krai",
    "russia-kurgan-oblast": "Kurgan Oblast",
    "russia-kursk-oblast": "Kursk Oblast",
    "russia-leningrad-oblast": "Leningrad Oblast",
    "russia-lipetsk-oblast": "Lipetsk Oblast",
    "russia-magadan-oblast": "Magadan Oblast",
    "russia-mari-el": "Mari El Republic",
    "russia-mordovia": "Republic of Mordovia",
    "russia-moscow": "Moscow",
    "russia-moscow-oblast": "Moscow Oblast",
    "russia-murmansk-oblast": "Murmansk Oblast",
    "russia-nenets": "Nenets Autonomous Okrug",
    "russia-nizhny-novgorod-oblast": "Nizhny Novgorod Oblast",
    "russia-north-ossetia": "North Ossetia–Alania",
    "russia-novgorod-oblast": "Novgorod Oblast",
    "russia-novosibirsk-oblast": "Novosibirsk Oblast",
    "russia-omsk-oblast": "Omsk Oblast",
    "russia-orenburg-oblast": "Orenburg Oblast",
    "russia-oryol-oblast": "Oryol Oblast",
    "russia-penza-oblast": "Penza Oblast",
    "russia-perm-krai": "Perm Krai",
    "russia-primorsky-krai": "Primorsky Krai",
    "russia-pskov-oblast": "Pskov Oblast",
    "russia-rostov-oblast": "Rostov Oblast",
    "russia-ryazan-oblast": "Ryazan Oblast",
    "russia-saint-petersburg": "Saint Petersburg",
    "russia-sakha": "Sakha (Yakutia)",
    "russia-sakhalin-oblast": "Sakhalin Oblast",
    "russia-samara-oblast": "Samara Oblast",
    "russia-saratov-oblast": "Saratov Oblast",
    "russia-sevastopol": "Sevastopol",
    "russia-smolensk-oblast": "Smolensk Oblast",
    "russia-stavropol-krai": "Stavropol Krai",
    "russia-sverdlovsk-oblast": "Sverdlovsk Oblast",
    "russia-tambov-oblast": "Tambov Oblast",
    "russia-tatarstan": "Republic of Tatarstan",
    "russia-tomsk-oblast": "Tomsk Oblast",
    "russia-tula-oblast": "Tula Oblast",
    "russia-tuva": "Tuva Republic",
    "russia-tver-oblast": "Tver Oblast",
    "russia-tyumen-oblast": "Tyumen Oblast",
    "russia-udmurtia": "Udmurt Republic",
    "russia-ulyanovsk-oblast": "Ulyanovsk Oblast",
    "russia-vladimir-oblast": "Vladimir Oblast",
    "russia-volgograd-oblast": "Volgograd Oblast",
    "russia-vologda-oblast": "Vologda Oblast",
    "russia-voronezh-oblast": "Voronezh Oblast",
    "russia-yamalo-nenets": "Yamalo-Nenets Autonomous Okrug",
    "russia-yaroslavl-oblast": "Yaroslavl Oblast",
    "russia-zabaykalsky-krai": "Zabaykalsky Krai",
}


def _num(v) -> float:
    if v in (None, "-", "–", "—", "...", "…"):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def _head(label: str) -> str:
    return label.split("(")[0].strip().lower()


def classify(label: str) -> str | None:
    head = _head(label)
    if head.startswith("все население") or head.startswith("указавшие национальную"):
        return None
    if head.startswith("лица, в переписных") or head.startswith("отказавшиеся"):
        return None
    if head in {"русские", "казаки", "поморы"}:
        return "russian"
    if head == "татары":
        return "tatar"
    if head.startswith("чеченцы"):
        return "chechen"
    if head.startswith("башкиры"):
        return "bashkir"
    return None


def row_label_value(cells) -> tuple[str | None, float | None]:
    """Some regional sheets prefix an MDX id in column A and shift counts right."""
    label = None
    for c in cells:
        if c is None:
            continue
        if isinstance(c, str):
            s = c.strip()
            if not s or s.startswith("["):
                continue
            if label is None:
                label = s
        elif isinstance(c, (int, float)) and label is not None:
            return label, float(c)
        elif label is not None:
            n = _num(c)
            if n:
                return label, n
    return None, None


def parse_sheet(ws) -> tuple[int, dict[str, float]] | None:
    total = 0.0
    declared = 0.0
    buckets = {"russian": 0.0, "tatar": 0.0, "chechen": 0.0, "bashkir": 0.0}
    for row in ws.iter_rows(max_col=6, values_only=True):
        label, val = row_label_value(row)
        if not label or val is None:
            continue
        head = _head(label)
        if head.startswith("все население"):
            total = val
            continue
        if head.startswith("указавшие национальную"):
            declared = val
            continue
        kind = classify(label)
        if kind:
            buckets[kind] += val
    if total <= 0:
        return None
    not_stated = max(0.0, total - declared)
    featured = sum(buckets.values())
    other = max(0.0, declared - featured)
    pct = lambda n: round(100.0 * n / total, 2)
    return int(round(total)), {
        "russian": pct(buckets["russian"]),
        "tatar": pct(buckets["tatar"]),
        "chechen": pct(buckets["chechen"]),
        "bashkir": pct(buckets["bashkir"]),
        "other": pct(other),
        "not_stated": pct(not_stated),
    }


def ensure_xlsx() -> Path:
    if XLSX_PATH.exists() and XLSX_PATH.stat().st_size > 100_000:
        return XLSX_PATH
    CACHE.mkdir(parents=True, exist_ok=True)
    print(f"  GET Rosstat Table 1…")
    req = urllib.request.Request(
        XLSX_URL, headers={"User-Agent": "birthrate.io census-maps/1.0"}
    )
    with urllib.request.urlopen(req, timeout=120) as r:
        XLSX_PATH.write_bytes(r.read())
    return XLSX_PATH


def catalog_entry() -> dict:
    return {
        "slug": "russia",
        "iso3": "RUS",
        "iso2": "RU",
        "name": "Russia",
        "kicker": "Census 2021",
        "title": "Ethnic group",
        "year": 2021,
        "source": SOURCE,
        "sourceUrl": SOURCE_URL,
        "nationalLabel": "Russia",
        "topicLabel": "Ethnic group",
        "groups": GROUPS,
        "levels": [
            {
                "id": "subject",
                "label": "Subject",
                "kind": "Federal subjects",
                "geoUrl": "/geo/census/rus-subject.json",
            }
        ],
        "dataUrl": "/data/census/rus.json",
        "fitMaxZoom": 4.4,
    }


def write_geo() -> None:
    geo = json.loads(GEO_SRC.read_text())
    for feat in geo["features"]:
        slug = (feat.get("properties") or {}).get("slug")
        if slug:
            feat["id"] = slug
            feat["properties"]["name"] = DISPLAY_NAMES.get(
                slug, feat["properties"].get("name", slug)
            )
    OUT_GEO.mkdir(parents=True, exist_ok=True)
    dest = OUT_GEO / "rus-subject.json"
    dest.write_text(json.dumps(geo, ensure_ascii=False, separators=(",", ":")))
    print(f"  wrote {dest.relative_to(ROOT)} ({len(geo['features'])} subjects)")


def upsert_catalog(entry: dict) -> None:
    if OUT_CATALOG.exists():
        payload = json.loads(OUT_CATALOG.read_text())
        countries = [c for c in payload.get("countries", []) if c.get("slug") != "russia"]
    else:
        countries = []
    featured = ["denmark", "germany", "spain", "russia", "france", "italy", "uk"]
    countries.append(entry)
    countries.sort(
        key=lambda c: (
            featured.index(c["slug"]) if c["slug"] in featured else 50,
            c.get("name", ""),
        )
    )
    OUT_CATALOG.parent.mkdir(parents=True, exist_ok=True)
    OUT_CATALOG.write_text(
        json.dumps({"countries": countries}, ensure_ascii=False, indent=2) + "\n"
    )


def build_russia() -> dict:
    import openpyxl

    ensure_xlsx()
    write_geo()
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    national = parse_sheet(wb["Российская Федерация"])
    if not national:
        raise SystemExit("Russia national sheet had no total")
    nat_pop, nat_shares = national
    areas: dict[str, dict] = {}
    missing_sheets = []
    for sheet, slug in SHEET_TO_SLUG.items():
        if sheet not in wb.sheetnames:
            missing_sheets.append(sheet)
            continue
        parsed = parse_sheet(wb[sheet])
        if not parsed:
            print(f"  skip {sheet}: no total")
            continue
        pop, shares = parsed
        areas[slug] = {
            "code": slug,
            "name": DISPLAY_NAMES.get(slug, slug),
            "slug": slug,
            "population": pop,
            "shares": shares,
        }
    wb.close()
    if missing_sheets:
        raise SystemExit(f"Missing sheets: {missing_sheets}")
    expected = set(SHEET_TO_SLUG.values())
    if set(areas) != expected:
        raise SystemExit(
            f"Area mismatch extra={set(areas)-expected} missing={expected-set(areas)}"
        )
    payload = {
        "source": SOURCE,
        "sourceUrl": SOURCE_URL,
        "year": 2021,
        "unit": "%",
        "national": {"population": nat_pop, "shares": nat_shares},
        "areas": {"subject": areas},
    }
    OUT_DATA.mkdir(parents=True, exist_ok=True)
    dest = OUT_DATA / "rus.json"
    dest.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")))
    print(
        f"  wrote {dest.relative_to(ROOT)} "
        f"national Russians {nat_shares['russian']}%  ({len(areas)} subjects)"
    )
    entry = catalog_entry()
    return entry


def main() -> None:
    entry = build_russia()
    upsert_catalog(entry)
    print("catalog upserted russia")


if __name__ == "__main__":
    sys.exit(main() or 0)
